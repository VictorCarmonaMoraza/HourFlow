import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from datetime import datetime
import os
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timedelta
from openpyxl.styles import PatternFill, Font


def mostrar_ventana_proyectos(nombre_usuario):
    ventana_proyectos = tk.Tk()
    ventana_proyectos.title("Proyectos")

    ventana_proyectos.minsize(200, 100)

    etiqueta_bienvenida = tk.Label(ventana_proyectos, text=f"¡Bienvenido, {nombre_usuario}!")
    etiqueta_bienvenida.pack()

    file_path = None
    df = None
    proyectos = []
    datos_proyectos = []  # Lista para almacenar los datos de los proyectos
    tiempo_inicio = None
    tiempo_fin = None
    proyecto_anterior = None
    contador_pausado = False  # Variable para rastrear si el contador está pausado
    color_contador = "black"  # Color inicial del contador
    tiempo_pausa_inicio = None  # Variable para almacenar el tiempo de inicio de la pausa
    tiempo_acumulado_pausa = 0  # Variable para almacenar el tiempo acumulado durante la pausa

    def cargar_excel():
        nonlocal file_path, df, proyectos
        file_path = filedialog.askopenfilename(filetypes=[("Archivos Excel", "*.xlsx")])
        if file_path:
            try:
                df = pd.read_excel(file_path, sheet_name="Hoja1")
                df.rename(columns={"CÓDIGO PROYECTO": "Código Proyecto", "DESCRIPCIÓN": "Descripción"}, inplace=True)

                proyectos = df["Código Proyecto"].tolist()

                variable_proyecto.set(proyectos[0])
                desplegable_proyecto["menu"].delete(0, "end")
                for proyecto in proyectos:
                    desplegable_proyecto["menu"].add_command(label=proyecto,
                                                             command=lambda p=proyecto: cambiar_proyecto(p, nombre_usuario))

            except FileNotFoundError:
                messagebox.showerror("Error", "Archivo no encontrado.")
            except pd.errors.ParserError:
                messagebox.showerror("Error", "Error al leer el archivo Excel.")
            except Exception as e:
                messagebox.showerror("Error", f"Error inesperado: {e}")

    boton_cargar = tk.Button(ventana_proyectos, text="Cargar Excel", command=cargar_excel)
    boton_cargar.pack()

    frame_proyecto = tk.Frame(ventana_proyectos)
    frame_proyecto.pack()

    variable_proyecto = tk.StringVar(ventana_proyectos)
    variable_proyecto.set("Seleccionar proyecto")
    desplegable_proyecto = tk.OptionMenu(frame_proyecto, variable_proyecto, "Seleccionar proyecto")
    desplegable_proyecto.pack(side=tk.LEFT)

    entrada_buscador = tk.Entry(frame_proyecto)
    entrada_buscador.pack(side=tk.LEFT)

    def buscar_proyecto(*args):
        texto_busqueda = entrada_buscador.get().lower()
        desplegable_proyecto["menu"].delete(0, "end")
        proyectos_filtrados = [proyecto for proyecto in proyectos if texto_busqueda in proyecto.lower()]
        for proyecto in proyectos_filtrados:
            desplegable_proyecto["menu"].add_command(label=proyecto,
                                                     command=lambda p=proyecto: cambiar_proyecto(p, nombre_usuario))
        if proyectos_filtrados:
            variable_proyecto.set(proyectos_filtrados[0])
        else:
            variable_proyecto.set("Seleccionar proyecto")

    entrada_buscador.bind("<KeyRelease>", buscar_proyecto)

    etiqueta_contador = tk.Label(ventana_proyectos, text="00:00:00", font=("Helvetica", 48), fg=color_contador)
    etiqueta_contador.pack()

    def cambiar_proyecto(proyecto_seleccionado, nombre_usuario):
        nonlocal tiempo_inicio, tiempo_fin, proyecto_anterior, df, file_path, datos_proyectos, contador_pausado, color_contador
        tiempo_fin = datetime.now()

        if tiempo_inicio and proyecto_anterior and file_path is not None:
            try:
                indice = proyectos.index(proyecto_anterior)

                # Calcular el tiempo invertido
                tiempo_invertido = tiempo_fin - tiempo_inicio
                horas, segundos_totales = divmod(tiempo_invertido.seconds, 3600)
                minutos, segundos = divmod(segundos_totales, 60)

                # Agregar los datos del proyecto a la lista, incluyendo la descripción
                datos_proyectos.append({
                    "Código Proyecto": proyecto_anterior,
                    "Descripción": df.at[indice, "Descripción"],  # Agregar la descripción
                    "Usuario": nombre_usuario,
                    "Start": tiempo_inicio.strftime("%Y-%m-%d %H:%M:%S"),
                    "End": tiempo_fin.strftime("%Y-%m-%d %H:%M:%S"),
                    "Tiempo Invertido": f"{horas:02d}:{minutos:02d}:{segundos:02d}"
                })

            except Exception as e:
                messagebox.showerror("Error", f"Error al actualizar Excel: {e}")

        tiempo_inicio = datetime.now()
        print(f"tiempo_inicio: {tiempo_inicio}")
        proyecto_anterior = proyecto_seleccionado
        variable_proyecto.set(proyecto_seleccionado)
        contador_pausado = False  # Reiniciar el estado de pausa
        color_contador = "black"  # Reiniciar el color del contador
        etiqueta_contador.config(fg=color_contador)  # Actualizar el color del contador
        actualizar_contador()

    def actualizar_contador():
        nonlocal tiempo_inicio, contador_pausado, color_contador, tiempo_acumulado_pausa
        print("actualizar_contador() llamada")
        if tiempo_inicio and not contador_pausado:
            tiempo_transcurrido = datetime.now() - tiempo_inicio
            tiempo_total = tiempo_transcurrido.seconds + tiempo_acumulado_pausa
            horas, segundos_totales = divmod(tiempo_total, 3600)
            minutos, segundos = divmod(segundos_totales, 60)
            print(f"tiempo_transcurrido: {tiempo_transcurrido}")
            etiqueta_contador.config(text=f"{horas:02d}:{minutos:02d}:{segundos:02d}")
            ventana_proyectos.after(1000, actualizar_contador)
        elif contador_pausado and tiempo_pausa_inicio:
            tiempo_transcurrido_pausa = datetime.now() - tiempo_pausa_inicio
            horas_pausa, segundos_totales_pausa = divmod(tiempo_transcurrido_pausa.seconds, 3600)
            minutos_pausa, segundos_pausa = divmod(segundos_totales_pausa, 60)
            etiqueta_contador.config(text=f"{horas_pausa:02d}:{minutos_pausa:02d}:{segundos_pausa:02d}")
            ventana_proyectos.after(1000, actualizar_contador)

    def pausar_contador():
        nonlocal contador_pausado, color_contador, tiempo_inicio, proyecto_anterior, datos_proyectos, df, nombre_usuario, tiempo_pausa_inicio, tiempo_acumulado_pausa

        contador_pausado = not contador_pausado  # Cambiar el estado de pausa
        if contador_pausado:
            color_contador = "green"  # Cambiar el color del contador a verde
            tiempo_pausa_inicio = datetime.now()  # Guardar el tiempo actual para calcular el tiempo transcurrido
            if tiempo_inicio:
                tiempo_fin = datetime.now()
                tiempo_transcurrido = tiempo_fin - tiempo_inicio
                horas, segundos_totales = divmod(tiempo_transcurrido.seconds, 3600)
                minutos, segundos = divmod(segundos_totales, 60)

                # Agregar los datos del proyecto a la lista, incluyendo la descripción
                datos_proyectos.append({
                    "Código Proyecto": proyecto_anterior,
                    "Descripción": df.at[proyectos.index(proyecto_anterior), "Descripción"],  # Agregar la descripción
                    "Usuario": "Prueba Pausa Victor",
                    "Start": tiempo_inicio.strftime("%Y-%m-%d %H:%M:%S"),
                    "End": tiempo_fin.strftime("%Y-%m-%d %H:%M:%S"),
                    "Tiempo Invertido": f"{horas:02d}:{minutos:02d}:{segundos:02d}"
                })

                # Cambiar el proyecto actual al proyecto especificado (23-BONP-NOP-0097-00)
                proyecto_anterior = "23-BONP-NOP-0097-00"
                variable_proyecto.set(proyecto_anterior)

                # No reiniciar tiempo_inicio, solo guardar el tiempo transcurrido
                tiempo_acumulado_pausa = tiempo_transcurrido.seconds
            else:
                tiempo_acumulado_pausa = 0

            etiqueta_contador.config(
                text=f"{horas:02d}:{minutos:02d}:{segundos:02d}")  # Mostrar el tiempo transcurrido en el contador
            etiqueta_contador.config(fg=color_contador)  # Actualizar el color del contador

        else:
            color_contador = "black"  # Cambiar el color del contador a negro
            etiqueta_contador.config(fg=color_contador)  # Actualizar el color del contador
            tiempo_inicio = datetime.now()  # Reiniciar el tiempo de inicio al reanudar
            if tiempo_pausa_inicio:
                tiempo_acumulado_pausa += (datetime.now() - tiempo_pausa_inicio).seconds
            tiempo_pausa_inicio = None  # Reiniciar el tiempo de pausa
            actualizar_contador()  # Reanudar el contador

    def finalizar_proceso():
        nonlocal tiempo_inicio, tiempo_fin, proyecto_anterior, df, file_path, datos_proyectos
        if messagebox.askyesno("Confirmar", "¿Estás seguro de que quieres finalizar el proceso?"):
            print(f"tiempo_inicio: {tiempo_inicio}")
            print(f"proyecto_anterior: {proyecto_anterior}")
            print(f"file_path: {file_path}")
            if tiempo_inicio and proyecto_anterior and file_path is not None:
                try:
                    print(f"Datos a escribir: {datos_proyectos}")
                    print(f"Ruta del archivo: {file_path}")
                    tiempo_fin = datetime.now()
                    tiempo_invertido = tiempo_fin - tiempo_inicio
                    horas, segundos_totales = divmod(tiempo_invertido.seconds, 3600)
                    minutos, segundos = divmod(segundos_totales, 60)
                    datos_proyectos.append({
                        "Código Proyecto": proyecto_anterior,
                        "Descripción": df.at[proyectos.index(proyecto_anterior), "Descripción"],
                        "Usuario": nombre_usuario,
                        "Start": tiempo_inicio.strftime("%Y-%m-%d %H:%M:%S"),
                        "End": tiempo_fin.strftime("%Y-%m-%d %H:%M:%S"),
                        "Tiempo Invertido": f"{horas:02d}:{minutos:02d}:{segundos:02d}"
                    })
                    crear_o_actualizar_pestaña(file_path, datos_proyectos)
                except Exception as e:
                    messagebox.showerror("Error", f"Error al actualizar Excel: {e}")
            ventana_proyectos.destroy()

    def comprobar_pestaña_fecha_actual(nombre_archivo):
        """
        Comprueba si existe una pestaña con la fecha actual en un archivo Excel.

        Args:
            nombre_archivo (str): La ruta al archivo Excel.

        Returns:
            bool: True si la pestaña existe, False en caso contrario.
        """
        try:
            # Cargar el archivo Excel
            libro_trabajo = openpyxl.load_workbook(nombre_archivo)

            # Obtener la fecha actual en formato "YYYY-MM-DD"
            fecha_actual = datetime.now().strftime("%Y-%m-%d")

            # Comprobar si la pestaña con la fecha actual existe
            if fecha_actual in libro_trabajo.sheetnames:
                return True
            else:
                return False

        except FileNotFoundError:
            print(f"Error: El archivo '{nombre_archivo}' no fue encontrado.")
            return False
        except Exception as e:
            print(f"Error inesperado: {e}")
            return False

    import pandas as pd
    import openpyxl
    from datetime import datetime, timedelta

    import pandas as pd
    import openpyxl
    from datetime import datetime, timedelta
    from openpyxl.utils.dataframe import dataframe_to_rows

    import pandas as pd
    import openpyxl
    from datetime import datetime, timedelta
    from openpyxl.utils.dataframe import dataframe_to_rows

    import pandas as pd
    import openpyxl
    from datetime import datetime, timedelta
    from openpyxl.utils.dataframe import dataframe_to_rows

    import pandas as pd
    import openpyxl
    from datetime import datetime, timedelta
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import PatternFill, Font

    def crear_o_actualizar_pestaña(nombre_archivo, datos):
        """
        Crea una nueva pestaña con la fecha actual o actualiza la existente, sumando tiempos si el proyecto ya existe.

        Args:
            nombre_archivo (str): La ruta al archivo Excel.
            datos (list): Una lista de diccionarios con los datos a escribir en la pestaña.
        """
        try:
            # Cargar el archivo Excel
            libro_trabajo = openpyxl.load_workbook(nombre_archivo)

            # Obtener la fecha actual en formato "YYYY-MM-DD"
            fecha_actual = datetime.now().strftime("%Y-%m-%d")

            # Comprobar si la pestaña con la fecha actual existe
            if fecha_actual in libro_trabajo.sheetnames:
                # Seleccionar la pestaña existente
                hoja = libro_trabajo[fecha_actual]
            else:
                # Crear una nueva pestaña con la fecha actual
                hoja = libro_trabajo.create_sheet(fecha_actual)

            # Leer los datos existentes de la pestaña en un DataFrame
            df_existente = pd.DataFrame(hoja.values)
            if df_existente.empty:
                df_existente = pd.DataFrame(
                    columns=["Código Proyecto", "Descripción", "Usuario", "Start", "End", "Tiempo Invertido"])
            else:
                df_existente.columns = df_existente.iloc[0]
                df_existente = df_existente[1:]

                # Convertir la columna "Tiempo Invertido" a timedelta para sumar tiempos
                df_existente["Tiempo Invertido"] = pd.to_timedelta(df_existente["Tiempo Invertido"], errors='coerce')

            # Convertir los nuevos datos a un DataFrame
            df_nuevo = pd.DataFrame(datos)

            # Convertir la columna "Tiempo Invertido" a timedelta para sumar tiempos
            df_nuevo["Tiempo Invertido"] = pd.to_timedelta(df_nuevo["Tiempo Invertido"], errors='coerce')

            # Verificar si existe un registro con el mismo código de proyecto
            for index, row in df_nuevo.iterrows():
                codigo_proyecto = row["Código Proyecto"]
                if codigo_proyecto in df_existente["Código Proyecto"].values:
                    # Sumar los tiempos invertidos
                    try:
                        tiempo_invertido_existente = \
                        df_existente.loc[df_existente["Código Proyecto"] == codigo_proyecto, "Tiempo Invertido"].iloc[0]
                        tiempo_invertido_nuevo = row["Tiempo Invertido"]
                        tiempo_invertido_sumado = tiempo_invertido_existente + tiempo_invertido_nuevo

                        # Actualizar el registro existente
                        df_existente.loc[df_existente[
                                             "Código Proyecto"] == codigo_proyecto, "Tiempo Invertido"] = tiempo_invertido_sumado
                    except IndexError:
                        # Si no se encuentra el registro, agregar el nuevo registro
                        if not df_existente.empty:  # Agregamos esta comprobación
                            df_existente = pd.concat([df_existente, pd.DataFrame([row])], ignore_index=True)
                        else:
                            df_existente = pd.DataFrame([row])
                else:
                    # Agregar el nuevo registro al DataFrame existente
                    if not df_existente.empty:  # Agregamos esta comprobación
                        df_existente = pd.concat([df_existente, pd.DataFrame([row])], ignore_index=True)
                    else:
                        df_existente = pd.DataFrame([row])

            # Convertir la columna "Tiempo Invertido" a string para escribir en Excel
            df_existente["Tiempo Invertido"] = df_existente["Tiempo Invertido"].apply(
                lambda x: str(x).split(' ')[-1] if isinstance(x, timedelta) else str(x))

            # Escribir los datos actualizados en la pestaña
            hoja.delete_rows(1, hoja.max_row)  # Limpiar la hoja existente
            for r_idx, r in enumerate(dataframe_to_rows(df_existente, index=False, header=True)):
                hoja.append(r)
                if r_idx == 0:  # Aplicar estilo a la fila de cabeceras
                    colores = ["fc02b0", "3aa91c", "17f5eb", "E0E0E0", "F0F0F0", "FFFFFF"]  # Lista de colores
                    for c_idx, cell in enumerate(hoja[1]):
                        cell.fill = PatternFill(start_color=colores[c_idx % len(colores)],
                                                end_color=colores[c_idx % len(colores)],
                                                fill_type="solid")  # Color de relleno
                        cell.font = Font(color="000000")  # Color de fuente negro

            # Guardar el archivo Excel
            libro_trabajo.save(nombre_archivo)

        except FileNotFoundError:
            print(f"Error: El archivo '{nombre_archivo}' no fue encontrado.")
        except Exception as e:
            print(f"Error inesperado: {e}")
    def cerrar_ventana():
        nonlocal tiempo_inicio, tiempo_fin, proyecto_anterior, df, file_path, datos_proyectos
        if tiempo_inicio and proyecto_anterior and file_path is not None:
            try:
                tiempo_fin = datetime.now()
                tiempo_invertido = tiempo_fin - tiempo_inicio
                horas, segundos_totales = divmod(tiempo_invertido.seconds, 3600)
                minutos, segundos = divmod(segundos_totales, 60)
                datos_proyectos.append({
                    "Código Proyecto": proyecto_anterior,
                    "Descripción": df.at[proyectos.index(proyecto_anterior), "Descripción"],
                    "Usuario": nombre_usuario,
                    "Start": tiempo_inicio.strftime("%Y-%m-%d %H:%M:%S"),
                    "End": tiempo_fin.strftime("%Y-%m-%d %H:%M:%S"),
                    "Tiempo Invertido": f"{horas:02d}:{minutos:02d}:{segundos:02d}"
                })
                crear_o_actualizar_pestaña(file_path, datos_proyectos)
            except Exception as e:
                messagebox.showerror("Error", f"Error al cerrar la ventana: {e}")
        ventana_proyectos.destroy()

    boton_pausar = tk.Button(ventana_proyectos, text="Pausar", command=pausar_contador)
    boton_pausar.pack()

    boton_finalizar = tk.Button(ventana_proyectos, text="Finalizar", command=finalizar_proceso)
    boton_finalizar.pack()

    ventana_proyectos.protocol("WM_DELETE_WINDOW", cerrar_ventana)

    ventana_proyectos.mainloop()